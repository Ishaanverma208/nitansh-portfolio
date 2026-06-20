"""
Generate preview images from Excel files for the portfolio website.
Finds the densest data region across all sheets and renders it as a styled dark-themed PNG.
"""
import os
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

EXCEL_FILES = {
    "polycab": "Polycab India Model final.xlsx",
    "havells": "Havells final model.xlsx",
    "maruti": "Maruti Suzuki Final Analysis.xlsx",
    "hyundai": "Hyundai Motors India Final Model.xlsx",
}

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "public", "images")
PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))

# Dark theme colors
BG_COLOR = (10, 15, 30)
HEADER_BG = (20, 25, 38)
CELL_BG_ALT = (14, 19, 34)
GRID_COLOR = (30, 42, 65)
TEXT_COLOR = (190, 200, 220)
HEADER_TEXT = (245, 158, 11)
NUMBER_COLOR = (100, 170, 255)
LABEL_COLOR = (160, 180, 210)
NEGATIVE_COLOR = (239, 100, 100)

IMG_WIDTH = 1200
IMG_HEIGHT = 700
ROW_HEIGHT = 28
HEADER_HEIGHT = 34
PADDING_LEFT = 10
MIN_COL_WIDTH = 100
FIRST_COL_WIDTH = 200


def get_font(size=12, bold=False):
    """Try to load a monospace font."""
    if bold:
        candidates = ["consolab.ttf", "courbd.ttf", "arialbd.ttf", "Consolab.ttf"]
    else:
        candidates = ["consola.ttf", "cour.ttf", "arial.ttf", "Consola.ttf"]
    for name in candidates:
        try:
            return ImageFont.truetype(name, size)
        except (OSError, IOError):
            continue
    return ImageFont.load_default()


def find_best_sheet(wb):
    """Find the sheet with the most non-empty cells."""
    best_sheet = None
    best_score = 0

    for ws in wb.worksheets:
        score = 0
        try:
            for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row or 1),
                                     max_col=min(12, ws.max_column or 1), values_only=True):
                for cell in row:
                    if cell is not None and str(cell).strip():
                        score += 1
        except Exception:
            continue

        if score > best_score:
            best_score = score
            best_sheet = ws

    return best_sheet


def find_dense_region(ws, max_rows=25, max_cols=8):
    """Find the starting row/col with the densest data block."""
    all_data = []
    scan_rows = min(100, ws.max_row or 1)
    scan_cols = min(15, ws.max_column or 1)

    for row in ws.iter_rows(min_row=1, max_row=scan_rows, max_col=scan_cols, values_only=True):
        all_data.append([str(cell) if cell is not None else "" for cell in row])

    # Score each possible starting position
    best_score = 0
    best_start_row = 0
    best_start_col = 0

    for start_row in range(min(30, len(all_data))):
        for start_col in range(min(8, scan_cols)):
            score = 0
            for r in range(start_row, min(start_row + max_rows, len(all_data))):
                for c in range(start_col, min(start_col + max_cols, len(all_data[r]) if r < len(all_data) else 0)):
                    if all_data[r][c].strip():
                        score += 1
            if score > best_score:
                best_score = score
                best_start_row = start_row
                best_start_col = start_col

    # Extract the dense region
    result = []
    for r in range(best_start_row, min(best_start_row + max_rows, len(all_data))):
        row_data = []
        for c in range(best_start_col, min(best_start_col + max_cols, len(all_data[r]) if r < len(all_data) else 0)):
            row_data.append(all_data[r][c])
        result.append(row_data)

    return result


def format_cell_value(value):
    """Format cell values for display."""
    if not value or value == "None":
        return ""
    try:
        num = float(value)
        if abs(num) >= 1e6:
            return f"{num/1e6:,.1f}M"
        elif abs(num) >= 1e3:
            return f"{num/1e3:,.1f}K"
        elif abs(num) < 0.01 and num != 0:
            return f"{num:.4f}"
        elif abs(num) < 1:
            return f"{num:.2%}" if abs(num) < 0.5 else f"{num:.2f}"
        elif num == int(num):
            return f"{int(num):,}"
        else:
            return f"{num:,.2f}"
    except (ValueError, TypeError):
        return str(value)[:28]


def render_data_to_image(data, output_path, company_name):
    """Render data grid as a dark-themed financial model PNG."""
    if not data or all(not any(cell.strip() for cell in row) for row in data):
        print(f"  No usable data for {company_name}")
        return False

    # Filter out completely empty rows
    data = [row for row in data if any(cell.strip() for cell in row)]
    if not data:
        return False

    num_cols = max(len(row) for row in data)
    # Ensure all rows have same number of columns
    data = [row + [""] * (num_cols - len(row)) for row in data]

    img = Image.new('RGB', (IMG_WIDTH, IMG_HEIGHT), BG_COLOR)
    draw = ImageDraw.Draw(img)

    font_regular = get_font(12)
    font_bold = get_font(12, bold=True)

    # Calculate column widths
    available_width = IMG_WIDTH - PADDING_LEFT * 2
    col_width = max(MIN_COL_WIDTH, available_width // num_cols)

    y = 6
    rows_drawn = 0

    for row_idx, row in enumerate(data):
        if y + ROW_HEIGHT > IMG_HEIGHT - 4:
            break

        is_header_row = row_idx == 0
        row_h = HEADER_HEIGHT if is_header_row else ROW_HEIGHT

        # Determine if this is a "section header" row (first cell has text, rest mostly empty)
        non_empty_count = sum(1 for cell in row if cell.strip())
        is_section = non_empty_count <= 2 and row[0].strip()

        # Row background
        if is_header_row:
            draw.rectangle([0, y, IMG_WIDTH, y + row_h], fill=HEADER_BG)
        elif is_section:
            draw.rectangle([0, y, IMG_WIDTH, y + row_h], fill=(18, 22, 38))
        elif row_idx % 2 == 0:
            draw.rectangle([0, y, IMG_WIDTH, y + row_h], fill=CELL_BG_ALT)

        # Grid lines
        draw.line([(0, y + row_h), (IMG_WIDTH, y + row_h)], fill=GRID_COLOR, width=1)

        x = PADDING_LEFT
        for col_idx, cell_value in enumerate(row):
            w = FIRST_COL_WIDTH if col_idx == 0 else col_width

            # Vertical grid
            if col_idx > 0:
                draw.line([(x, y), (x, y + row_h)], fill=GRID_COLOR, width=1)

            text = format_cell_value(cell_value)

            # Choose color & font
            if is_header_row:
                color = HEADER_TEXT
                f = font_bold
            elif is_section:
                color = LABEL_COLOR
                f = font_bold
            elif col_idx == 0:
                color = LABEL_COLOR
                f = font_regular
            else:
                # Check for negative numbers
                try:
                    num_val = float(cell_value.replace(',', ''))
                    color = NEGATIVE_COLOR if num_val < 0 else NUMBER_COLOR
                except (ValueError, TypeError):
                    color = TEXT_COLOR
                f = font_regular

            text_y = y + (row_h - 12) // 2
            # Right-align numbers, left-align text
            if col_idx > 0:
                try:
                    float(cell_value.replace(',', '').replace('%', ''))
                    # Right-align: calculate text width
                    bbox = draw.textbbox((0, 0), text, font=f)
                    text_w = bbox[2] - bbox[0]
                    draw.text((x + w - text_w - 8, text_y), text, fill=color, font=f)
                except (ValueError, TypeError):
                    draw.text((x + 6, text_y), text, fill=color, font=f)
            else:
                draw.text((x + 6, text_y), text, fill=color, font=f)

            x += w

        y += row_h
        rows_drawn += 1

    # Draw remaining vertical grid lines
    x = PADDING_LEFT + FIRST_COL_WIDTH
    for i in range(num_cols - 1):
        draw.line([(x, 0), (x, IMG_HEIGHT)], fill=GRID_COLOR, width=1)
        x += col_width

    # Add a subtle gradient at the bottom to fade out
    for fade_y in range(IMG_HEIGHT - 60, IMG_HEIGHT):
        alpha = int(255 * (fade_y - (IMG_HEIGHT - 60)) / 60)
        draw.line([(0, fade_y), (IMG_WIDTH, fade_y)],
                  fill=(BG_COLOR[0], BG_COLOR[1], BG_COLOR[2]), width=1)

    img.save(output_path, 'PNG', quality=95)
    print(f"  Saved: {output_path} ({rows_drawn} rows)")
    return True


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for company_id, filename in EXCEL_FILES.items():
        filepath = os.path.join(PROJECT_ROOT, filename)
        output_path = os.path.join(OUTPUT_DIR, f"{company_id}-preview.png")

        if not os.path.exists(filepath):
            print(f"  Warning: File not found: {filepath}")
            continue

        print(f"Processing: {filename}")

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            best_sheet = find_best_sheet(wb)
            if best_sheet is None:
                print(f"  No usable sheets found")
                wb.close()
                continue

            print(f"  Using sheet: '{best_sheet.title}'")
            data = find_dense_region(best_sheet, max_rows=25, max_cols=8)
            wb.close()

            render_data_to_image(data, output_path, company_id)

        except Exception as e:
            print(f"  Error: {e}")
            import traceback
            traceback.print_exc()

    print("\nDone! Preview images generated.")


if __name__ == "__main__":
    main()
